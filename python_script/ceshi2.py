#!/usr/bin/env python3
#coding=utf-8
import json
import datetime
'''
测试联盟牌不同品质对应的属性数量，战力，卡牌积分，分解材料
'''
from openpyxl import load_workbook
f_data=load_workbook(r"C:\Users\playcrab\Desktop\tarotCard.xlsx")
dict_data=[]
split_title_data=[]
new_rows=[]
for sheet in f_data.sheetnames:
    #只读取sheet1工作表内容，如果不加判断读取所有工作表中内容
    if sheet == "tarotCard":
        # 读取sheet1工作表的所有内容以列表形式存在sheet_data变量中
        sheet_data=list(f_data[sheet].iter_rows(min_row=1,min_col=1,values_only=True))
        for title in sheet_data[0]:
            #if title is not None:
            #split_title_data.append(title.split("$")[0])
            split_title_data.append(title)
        rows=sheet_data[5:]
        for i in rows:
                if any(i):
                    data_to_dict = dict(zip(split_title_data,i))
                    if not all(data_to_dict.keys()):
                        del data_to_dict[None]
                    dict_data.append(data_to_dict)
                else:
                    continue
for i in dict_data:
    if i["type$cs"] == 2 and i["color$cs"] == 3 :
        if i["basicAttNum$cs"] == 3:
            pass
        else:
            print("{0}的{1}字段配置不是3".format(i["id$cs"],i["basicAttNum$cs"]))
        if i["power$cs"] == 2000:
            pass
        else:
            print("{0}的{1}字段配置不是2000".format(i["id$cs"],i["power$cs"]))
        if i["points$cs"] == 800:
            pass
        else:
            print("{0}的{1}字段配置不是800".format(i["id$cs"],i["points$cs"]))
        if "commonSand" in i["decompose$cs"] and "0" in i["decompose$cs"] and "100" in i["decompose$cs"]:
            pass
        else:
            print("{0}的{1}字段配置错误".format(i["id$cs"],i["decompose$cs"]))
    if i["type$cs"] == 2 and i["color$cs"] == 4 :
        if i["basicAttNum$cs"] == 3:
            pass
        else:
            print("{0}的{1}字段配置不是3".format(i["id$cs"],i["basicAttNum$cs"]))
        if i["power$cs"] == 2500:
            pass
        else:
            print("{0}的{1}字段配置不是2500".format(i["id$cs"],i["power$cs"]))
        if i["points$cs"] == 1200:
            pass
        else:
            print("{0}的{1}字段配置不是1200".format(i["id$cs"],i["points$cs"]))
        if "legendSand" in i["decompose$cs"] and "0" in i["decompose$cs"] and "30" in i["decompose$cs"]:
            pass
        else:
            print("{0}的{1}字段配置错误".format(i["id$cs"],i["decompose$cs"]))
    if i["type$cs"] == 2 and i["color$cs"] == 5 :
        if i["basicAttNum$cs"] == 4:
            pass
        else:
            print("{0}的{1}字段配置不是4".format(i["id$cs"],i["basicAttNum$cs"]))
        if i["power$cs"] == 3000:
            pass
        else:
            print("{0}的{1}字段配置不是3000".format(i["id$cs"],i["power$cs"]))
        if i["points$cs"] == 1700:
            pass
        else:
            print("{0}的{1}字段配置不是1700".format(i["id$cs"],i["points$cs"]))
        if "legendSand" in i["decompose$cs"] and "0" in i["decompose$cs"] and "80" in i["decompose$cs"]:
            pass
        else:
            print("{0}的{1}字段配置错误".format(i["id$cs"],i["decompose$cs"]))