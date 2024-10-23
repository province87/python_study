#!/usr/bin/env python3
#coding=utf-8
import json
import datetime
from openpyxl import load_workbook
f_data=load_workbook(r"C:\Users\playcrab\Desktop\tarotCard.xlsx")
dict_data=[]
split_title_data=[]
new_rows=[]
for sheet in f_data.sheetnames:
    print(sheet)
    #只读取sheet1工作表内容，如果不加判断读取所有工作表中内容
    if sheet == "tarotCard":
        # 读取sheet1工作表的所有内容以列表形式存在sheet_data变量中
        sheet_data=list(f_data[sheet].iter_rows(min_row=1,min_col=1,values_only=True))
        print(sheet_data[0])
        for title in sheet_data[0]:
            #if title is not None:
            #split_title_data.append(title.split("$")[0])
            split_title_data.append(title)
        rows=sheet_data[5:]
        print(rows)
        for i in rows:
                if any(i):
                    data_to_dict = dict(zip(split_title_data,i))
                    if not all(data_to_dict.keys()):
                        del data_to_dict[None]
                    dict_data.append(data_to_dict)
                else:
                    continue
f_data.close()
print(dict_data)
dictdata_to_jsondata = json.dumps(dict_data,indent=2, ensure_ascii=False)
with open(r"C:\Users\playcrab\Desktop\tarotCard.json","w") as f:
    f.write(dictdata_to_jsondata)
print(dictdata_to_jsondata)